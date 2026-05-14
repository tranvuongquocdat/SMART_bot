#!/usr/bin/env python
"""Regenerate docs/self_test_clusters.xlsx from docs/self_test_clusters.csv.

Run after editing the CSV by hand or after a fresh `scripts/self_test.py`
sweep changes the status of any case. Produces a formatted workbook with
color-coded PASS/FAIL/NOT_TESTED, cluster highlights, wrapped cells and a
frozen header row.

Usage:
    .venv/bin/python scripts/gen_test_xlsx.py
"""
from __future__ import annotations

import csv
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SRC = Path("docs/self_test_clusters.csv")
DST = Path("docs/self_test_clusters.xlsx")

STATUS_FILL = {
    "PASS": PatternFill("solid", fgColor="C6EFCE"),
    "FAIL": PatternFill("solid", fgColor="FFC7CE"),
    "NOT_TESTED": PatternFill("solid", fgColor="FFEB9C"),
    "FEATURE_GAP": PatternFill("solid", fgColor="E7E6E6"),
}
STATUS_FONT = {
    "PASS": Font(color="006100", bold=True),
    "FAIL": Font(color="9C0006", bold=True),
    "NOT_TESTED": Font(color="9C6500", bold=True),
    "FEATURE_GAP": Font(color="595959", bold=True),
}
CLUSTER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMN_WIDTHS = {
    "A": 13,   # cluster
    "B": 36,   # cluster_purpose
    "C": 7,    # case_id
    "D": 42,   # case_name
    "E": 64,   # case_intent
    "F": 12,   # status
    "G": 36,   # evidence
    "H": 56,   # notes
}


def main() -> int:
    if not SRC.exists():
        print(f"error: {SRC} not found", file=sys.stderr)
        return 1

    with SRC.open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    if not rows:
        print(f"error: {SRC} is empty", file=sys.stderr)
        return 1

    headers, data = rows[0], rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Self-test scenarios"

    # ----- Summary block -----
    counts = Counter(r[5] for r in data)
    total = len(data)
    when = datetime.now().strftime("%Y-%m-%d %H:%M")
    pass_n = counts.get("PASS", 0)
    fail_n = counts.get("FAIL", 0)
    nottested_n = counts.get("NOT_TESTED", 0)
    gap_n = counts.get("FEATURE_GAP", 0)

    summary_title = Font(bold=True, color="FFFFFF", size=12)
    summary_fill = PatternFill("solid", fgColor="305496")
    ws.cell(row=1, column=1, value="SMART bot — self-test summary").font = summary_title
    ws.cell(row=1, column=1).fill = summary_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    info_rows = [
        ("Generated", when, None),
        ("Total cases", str(total), None),
        ("PASS", str(pass_n), "PASS"),
        ("FAIL", str(fail_n), "FAIL"),
        ("NOT_TESTED", str(nottested_n), "NOT_TESTED"),
        ("FEATURE_GAP", str(gap_n), "FEATURE_GAP"),
    ]
    for i, (label, val, status_key) in enumerate(info_rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=2, value=val).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=i, column=2).border = BORDER
        if status_key and status_key in STATUS_FILL:
            ws.cell(row=i, column=2).fill = STATUS_FILL[status_key]
            ws.cell(row=i, column=2).font = STATUS_FONT[status_key]

    # ----- Spacer row -----
    spacer_row = len(info_rows) + 2  # rows 1..N then blank
    ws.cell(row=spacer_row, column=1, value="")

    # ----- Header row -----
    header_row = spacer_row + 1
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[header_row].height = 24

    prev_cluster = None
    for r in data:
        cluster = r[0]
        if cluster != prev_cluster and prev_cluster is not None:
            ws.append([""] * len(headers))
        prev_cluster = cluster

        ws.append(r)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER

        ws.cell(row=row_idx, column=1).fill = CLUSTER_FILL
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=2).fill = CLUSTER_FILL

        status = r[5]
        if status in STATUS_FILL:
            sc = ws.cell(row=row_idx, column=6)
            sc.fill = STATUS_FILL[status]
            sc.font = STATUS_FONT[status]
            sc.alignment = Alignment(horizontal="center", vertical="center")

    for col, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = w
    # Freeze so the case header (and summary block above it) stays visible
    # while scrolling case rows.
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    wb.save(DST)
    print(f"Wrote {DST}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
